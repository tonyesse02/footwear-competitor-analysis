"""
scraper.py
==========
Core scraping classes for the Diadora Competitive Intelligence project.

Classes:
  - BrandScraper      : generic configurable scraper for a single brand
  - CompetitiveAnalyzer : orchestrator that runs all brand scrapers and
                          produces the final Excel report
"""

import re
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from config import BRAND_CONFIGS, CATEGORY_RULES


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def classify_product(name: str, url_category: str = "") -> Tuple[str, str]:
    """Classify a product name into (category, subcategory) using keyword rules."""
    text = (name + " " + url_category).lower()
    for category, rules in CATEGORY_RULES.items():
        if any(kw in text for kw in rules["keywords"]):
            subcategory = "Generale"
            for sub_name, sub_keywords in rules["subcategories"].items():
                if any(sk in text for sk in sub_keywords):
                    subcategory = sub_name
                    break
            return category, subcategory
    return "Altro", "Altro"


def build_driver(headless: bool = False) -> webdriver.Chrome:
    """Instantiate a Chrome WebDriver with anti-bot options."""
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1200")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(8)
    return driver


# ---------------------------------------------------------------------------
# BRAND SCRAPER
# ---------------------------------------------------------------------------

class BrandScraper:
    """Generic scraper configurable for any brand via BRAND_CONFIGS."""

    def __init__(self, brand_name: str, config: Dict, driver: webdriver.Chrome):
        self.brand_name = brand_name
        self.config     = config
        self.driver     = driver

    # ── Cookie handling ──────────────────────────────────────────────────────

    def handle_cookies(self) -> None:
        for by, selector in self.config["cookie_selectors"]:
            try:
                btn = WebDriverWait(self.driver, 4).until(
                    EC.element_to_be_clickable((by, selector))
                )
                btn.click()
                time.sleep(1)
                return
            except Exception:
                continue

    # ── Infinite scroll + load-more button ──────────────────────────────────

    def scroll_to_load(self, max_scrolls: int = 12) -> None:
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        for _ in range(max_scrolls):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2.5)
            try:
                btn = WebDriverWait(self.driver, 2).until(
                    EC.element_to_be_clickable(self.config["pagination_selector"])
                )
                btn.click()
                time.sleep(2)
            except Exception:
                pass
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    # ── CSS-pattern extraction helpers ──────────────────────────────────────

    def _text_by_class(self, container: BeautifulSoup, patterns: List) -> Optional[str]:
        for sel_type, pattern in patterns:
            if sel_type == "class":
                elem = container.find(class_=re.compile(pattern, re.I))
                if elem:
                    return elem.get_text(strip=True)
        return None

    def _extract_price(self, container: BeautifulSoup) -> Optional[float]:
        text = self._text_by_class(container, self.config["price_selectors"])
        if not text:
            return None
        matches = re.findall(r"\d{1,4}[,\.]\d{2}", text)
        if matches:
            try:
                return float(matches[0].replace(",", "."))
            except ValueError:
                pass
        matches = re.findall(r"\d{2,4}", text)
        if matches:
            try:
                return float(matches[0])
            except ValueError:
                pass
        return None

    def _extract_colors(self, container: BeautifulSoup) -> int:
        for sel_type, pattern in self.config["color_selectors"]:
            if sel_type == "class":
                elem = container.find(class_=re.compile(pattern, re.I))
                if elem:
                    match = re.search(r"(\d+)", elem.get_text())
                    if match:
                        return int(match.group(1))
                    swatches = elem.find_all(class_=re.compile(r"swatch|color-option", re.I))
                    if swatches:
                        return len(swatches)
        return 1

    def _extract_sizes(self, container: BeautifulSoup) -> List[str]:
        for sel_type, pattern in self.config.get("size_selectors", []):
            if sel_type == "class":
                elem = container.find(class_=re.compile(pattern, re.I))
                if elem:
                    sizes = elem.find_all(class_=re.compile(r"size|taglia", re.I))
                    return [s.get_text(strip=True) for s in sizes if s.get_text(strip=True)]
        return []

    def _find_containers(self, soup: BeautifulSoup) -> List:
        for tag, pattern in self.config["product_container"]:
            containers = soup.find_all(tag, class_=re.compile(pattern, re.I))
            if len(containers) >= 3:
                return containers
        return []

    # ── Scrape a single category page ────────────────────────────────────────

    def scrape_category(self, category: str, url: str, max_products: int = 100) -> List[Dict]:
        print(f"   [{self.brand_name}] {category} → {url}")
        try:
            self.driver.get(url)
            time.sleep(3)
            self.handle_cookies()
            self.scroll_to_load()

            soup       = BeautifulSoup(self.driver.page_source, "html.parser")
            containers = self._find_containers(soup)
            products   = []

            for container in containers[:max_products]:
                try:
                    name = self._text_by_class(container, self.config["name_selectors"])
                    if not name:
                        name = container.get_text(strip=True)[:80]

                    price       = self._extract_price(container)
                    colors      = self._extract_colors(container)
                    sizes       = self._extract_sizes(container)
                    cat, subcat = classify_product(name, category)

                    link        = container.find("a", href=True)
                    product_url = link["href"] if link else ""
                    if product_url and not product_url.startswith("http"):
                        product_url = self.config["base_url"].rstrip("/") + product_url

                    products.append({
                        "Brand":           self.brand_name,
                        "Modello":         name,
                        "Categoria URL":   category,
                        "Categoria":       cat,
                        "Sottocategoria":  subcat,
                        "Prezzo (€)":      price,
                        "N. Colori":       colors,
                        "Taglie":          ", ".join(sizes) if sizes else "N/D",
                        "N. Taglie":       len(sizes),
                        "URL Prodotto":    product_url,
                        "Data Rilevazione": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    })
                except Exception as e:
                    print(f"      ⚠ Errore prodotto: {e}")

            print(f"      ✓ {len(products)} prodotti estratti")
            return products

        except Exception as e:
            print(f"   ✗ Errore categoria {category}: {e}")
            return []

    def scrape_all(self, max_per_category: int = 80) -> List[Dict]:
        """Scrape all configured categories for this brand."""
        results = []
        for category, url in self.config["category_url_map"].items():
            results.extend(self.scrape_category(category, url, max_per_category))
            time.sleep(2)
        return results


# ---------------------------------------------------------------------------
# COMPETITIVE ANALYZER (ORCHESTRATOR)
# ---------------------------------------------------------------------------

class CompetitiveAnalyzer:
    """
    Orchestrates multi-brand scraping and produces the final Excel report.

    Usage:
        analyzer = CompetitiveAnalyzer(headless=False)
        df = analyzer.run(max_per_category=80)
        analyzer.build_report(df)
        analyzer.close()
    """

    def __init__(self, headless: bool = False, brands: Optional[List[str]] = None):
        self.brands   = brands or list(BRAND_CONFIGS.keys())
        self.driver   = build_driver(headless)
        self.all_data: List[Dict] = []

    def run(self, max_per_category: int = 80) -> pd.DataFrame:
        for brand_name in self.brands:
            if brand_name not in BRAND_CONFIGS:
                print(f"⚠ Brand '{brand_name}' not configured, skipping.")
                continue
            print(f"\n{'='*60}\n  {brand_name.upper()}\n{'='*60}")
            scraper      = BrandScraper(brand_name, BRAND_CONFIGS[brand_name], self.driver)
            brand_data   = scraper.scrape_all(max_per_category)
            self.all_data.extend(brand_data)
            print(f"  ✓ {brand_name}: {len(brand_data)} prodotti raccolti")

        df = pd.DataFrame(self.all_data)
        print(f"\nTotale prodotti raccolti: {len(df)}")
        return df

    def build_report(self, df: pd.DataFrame, output: str = "Diadora_Competitive_Analysis.xlsx") -> str:
        """Generate a formatted multi-sheet Excel report."""
        df["Prezzo (€)"] = pd.to_numeric(df["Prezzo (€)"], errors="coerce")

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer,                              sheet_name="Dati Grezzi",            index=False)
            df.groupby(["Brand", "Categoria"]).size().reset_index(name="N. Modelli")               .to_excel(writer,                              sheet_name="Modelli per Categoria",  index=False)
            df.groupby(["Brand", "Categoria", "Sottocategoria"])["Prezzo (€)"]               .agg(Media="mean", Minimo="min", Massimo="max", Mediana="median", N_Prodotti="count")               .round(2).reset_index()               .to_excel(writer,                              sheet_name="Analisi Prezzi",         index=False)
            df.pivot_table(index="Categoria", columns="Brand", values="Prezzo (€)", aggfunc="mean")               .round(2).to_excel(writer,                     sheet_name="Pivot Prezzi")

        self._format_report(output)
        print(f"✓ Report saved: {output}")
        return output

    def _format_report(self, filename: str) -> None:
        wb = openpyxl.load_workbook(filename)
        h_fill = PatternFill("solid", fgColor="1F3864")
        h_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
        a_fill = PatternFill("solid", fgColor="DCE6F1")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill, cell.font = h_fill, h_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                for cell in row:
                    cell.font = Font(name="Arial", size=10)
                    if r_idx % 2 == 0:
                        cell.fill = a_fill
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
        wb.save(filename)

    def close(self) -> None:
        if self.driver:
            self.driver.quit()
            print("✓ Browser closed")
